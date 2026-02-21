import pandas as pd
import numpy as np
import datetime
import re
import json
from dateutil.utils import today
from openpyxl import load_workbook
from datetime import timedelta

def load_config(config_path):
    """
    Load configuration from a JSON file.

    Opens the JSON file at the given path and returns its content as a dictionary.

    Parameters:
        config_path (str): Path to the JSON configuration file.

    Returns:
        dict: Parsed JSON configuration.
    """

    with open(config_path, "r") as f:
        return json.load(f)


def get_date_range(current_date):
    """
    Calculate the last complete week (Saturday to Friday) relative to the given date.

    Parameters:
        current_date (str or datetime.datetime): The reference date as a string in "YYYY-MM-DD" format or as a datetime object.

    Returns:
        tuple: A tuple containing two datetime objects:
            - The first element is the last Saturday (start of the week).
            - The second element is the last Friday (end of the week).

    Example:
        >>> last_saturday, last_friday = get_date_range("2023-04-10")
        >>> print(last_saturday, last_friday)
    """

    # Convert string date to datetime if needed
    if isinstance(current_date, str):
        current_date = datetime.datetime.strptime(current_date, "%Y-%m-%d")
    
    current_weekday = current_date.weekday()
    
    # Calculate days to last Friday (0=Mon, 1=Tue, ..., 6=Sun)
    days_to_last_friday = (current_weekday - 4) % 7
    if days_to_last_friday == 0:
        days_to_last_friday = 7
    
    last_friday = current_date - timedelta(days=days_to_last_friday)
    last_saturday = last_friday - timedelta(days=6)
    
    return last_saturday, last_friday

def calculate_week_uid(end_date):
    """
    Calculate a unique week identifier from the given end_date (Friday).

    The fiscal year starts on February 1 (adjusted to the first Saturday), and the week
    number is computed from this adjusted start date. The week identifier is returned
    as an integer in the format YYYYWW.

    Parameters:
        end_date (datetime.datetime): The end date (Friday) for the week.

    Returns:
        int: The week identifier.
    """

    year_start = datetime.datetime(
        end_date.year if end_date.month > 1 or end_date.day >= 1 else end_date.year - 1, 
        2, 
        1
    )
    
    # Adjust to first Saturday of the fiscal year
    if year_start.weekday() != 5:
        year_start += timedelta(days=(5 - year_start.weekday()) % 7)
    
    days_since_year_start = (end_date - year_start).days
    week_number = days_since_year_start // 7 + 1
    week_uid = f"{year_start.year}{week_number:02d}"
    
    return int(week_uid)

def preprocess_adset(text, patterns):
    """
    Classify adset names based on regex patterns.

    Converts the input text to lowercase and checks it against provided regex patterns.
    Returns the formatted category name if a match is found, otherwise returns "Unknown".

    Parameters:
        text (str): The adset name.
        patterns (dict): Dictionary with category names as keys and lists of regex patterns as values.

    Returns:
        str: The classified category name or "Unknown".
    """

    if not isinstance(text, str):
        return "Unknown"
        
    text = text.lower()
    
    for category, pattern_list in patterns.items():
        for pattern in pattern_list:
            if re.search(pattern, text):
                return category.replace("_", " ").title()
    
    return "Unknown"

def automate_excel_pivoting(config_file):
    """
    Main function to Automate Excel pivoting and update reports using configuration settings.

    Loads configuration from a JSON file, processes mobile installs and MAE spend reports,
    aggregates weekly data, and updates corresponding sheets in an Excel workbook.

    Parameters:
        config_file (str): Path to the JSON configuration file.
    """

    # Load configuration
    config = load_config(config_file)
    
    # Determine file paths and dates
    base_path = config["file_paths"]["base_path"]
    
    # Replace with current date instead of config["report_date"] for running it on real data
    current_date = config["report_date"]
    # current_date = datetime.datetime.today()
    
    # Use the fixed date from config
    current_date = datetime.datetime.strptime(current_date, "%Y-%m-%d")
    current_date_str = current_date.strftime("%m.%d.%Y")
    
    # Calculate date range for weekly report
    start_date, end_date = get_date_range(current_date)
    
    # Format filenames with current date
    mobile_installs_report = config["file_paths"]["mobile_installs_report"].format(date=current_date_str)
    mae_spend_report = config["file_paths"]["mae_spend_report"].format(date=end_date.strftime("%m.%d.%Y"))
    output_report = config["file_paths"]["output_report"]
    
    print(f"Processing data for week: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")
    
    # Load Mobile Installs Daily Spend Report
    spend_data = pd.read_excel(
        base_path + mobile_installs_report, 
        sheet_name="Sheet1"
    )
    
    # Load Mobile App MAE Daily Spend Report
    mae_spend_data = pd.read_excel(
        base_path + mae_spend_report, 
        sheet_name="Sheet1"
    )
    
    # Filter data for the current week
    mask = (spend_data['Date'] >= start_date) & (spend_data['Date'] <= end_date)
    filtered_spend_data = spend_data.loc[mask]
    
    # Calculate Week UID
    week_uid = calculate_week_uid(end_date)
    print(f"Week UID: {week_uid}")
    
    # Process MAI data
    filtered_mai_data = filtered_spend_data[filtered_spend_data['Type'] == 'MAI']
    mai_pivot_table = pd.pivot_table(
        filtered_mai_data, 
        values='Spend', 
        index=["Publisher"], 
        aggfunc="sum"
    ).reset_index()
    mai_pivot_table.set_index("Publisher", inplace=True)
    
    # Get publisher spend values
    publisher_spend = {}
    for publisher in config["sheets"]["channel_campaign_metrics"]["publisher_spend_mapping"].keys():
        try:
            publisher_spend[publisher] = mai_pivot_table.at[publisher, "Spend"]
        except KeyError:
            publisher_spend[publisher] = 0
            print(f"Warning: No data found for publisher '{publisher}'")
    
    # Process platform spend
    platform_spend = filtered_spend_data.groupby("Platform")["Spend"].sum().reset_index()
    platform_spend.set_index("Platform", inplace=True)
    
    platform_spend_values = {}
    for platform in config["sheets"]["overall_metrics"]["platform_spend_mapping"].keys():
        try:
            platform_spend_values[platform] = platform_spend.at[platform, "Spend"]
        except KeyError:
            platform_spend_values[platform] = 0
            print(f"Warning: No data found for platform '{platform}'")
    
    # Process MAE data
    mask = (mae_spend_data['Date'] >= start_date) & (mae_spend_data['Date'] <= end_date)
    filtered_mae_spend_data = mae_spend_data.loc[mask]
    
    # Preprocess adsets
    adset_patterns = {
        "Predictive_Churn": config["adset_patterns"]["predictive_churn"],
        "Low_Activity": config["adset_patterns"]["low_activity"]
    }
    
    filtered_mae_spend_data["Adset_cleaned"] = filtered_mae_spend_data["Adset"].apply(
        lambda x: preprocess_adset(x, adset_patterns) if pd.notnull(x) else "Unknown"
    )
    
    # Create Publisher-Platform combinations
    filtered_mae_spend_data["Publisher_Platform"] = (
        filtered_mae_spend_data["Publisher"] + "-" + filtered_mae_spend_data["Platform"]
    )
    
    # Aggregate by publisher-platform and adset
    platform_adset_spend = filtered_mae_spend_data.groupby(
        ["Publisher_Platform", "Adset_cleaned"]
    )["Spend"].sum().reset_index()
    
    # Convert to dictionary for easier lookup
    platform_adset_spend_dict = {}
    for _, row in platform_adset_spend.iterrows():
        key = f"{row['Publisher_Platform']}_{row['Adset_cleaned']}"
        platform_adset_spend_dict[key] = row['Spend']
    
    # Ensure all required keys exist
    for key in config["sheets"]["mae_audience_level"]["audience_mapping"].keys():
        if key not in platform_adset_spend_dict:
            platform_adset_spend_dict[key] = 0
    
    # Load and update the output Excel workbook
    print(f"Opening workbook: {base_path + output_report}")
    workbook = load_workbook(base_path + output_report)
    
    # Update Channel & Campaign Metrics sheet
    sheet_config = config["sheets"]["channel_campaign_metrics"]
    sheet = workbook[sheet_config["sheet_name"]]
    
    updated = False
    for row in range(sheet_config["start_row"], sheet.max_row + 1):
        if sheet[f"{sheet_config['week_uid_column']}{row}"].value == week_uid:
            for publisher, column in sheet_config["publisher_spend_mapping"].items():
                sheet[f"{column}{row}"] = publisher_spend.get(publisher, 0)
            updated = True
            print(f"Updated {sheet_config['sheet_name']} for week {week_uid}")
            break
    
    if not updated:
        print(f"Warning: Week {week_uid} not found in {sheet_config['sheet_name']}")
    
    # Update Overall Metrics sheet
    sheet_config = config["sheets"]["overall_metrics"]
    sheet = workbook[sheet_config["sheet_name"]]
    
    updated = False
    for row in range(sheet_config["start_row"], sheet.max_row + 1):
        if sheet[f"{sheet_config['week_uid_column']}{row}"].value == week_uid:
            for platform, column in sheet_config["platform_spend_mapping"].items():
                sheet[f"{column}{row}"] = platform_spend_values.get(platform, 0)
            updated = True
            print(f"Updated {sheet_config['sheet_name']} for week {week_uid}")
            break
    
    if not updated:
        print(f"Warning: Week {week_uid} not found in {sheet_config['sheet_name']}")
    
    # Update MAE Audience Level Data sheet
    sheet_config = config["sheets"]["mae_audience_level"]
    sheet = workbook[sheet_config["sheet_name"]]
    
    updated = False
    for row in range(sheet_config["start_row"], sheet.max_row + 1):
        if sheet[f"{sheet_config['week_uid_column']}{row}"].value == week_uid:
            for audience_key, column in sheet_config["audience_mapping"].items():
                sheet[f"{column}{row}"] = platform_adset_spend_dict.get(audience_key, 0)
            updated = True
            print(f"Updated {sheet_config['sheet_name']} for week {week_uid}")
            break
    
    if not updated:
        print(f"Warning: Week {week_uid} not found in {sheet_config['sheet_name']}")
    
    # Save the workbook
    print(f"Saving workbook: {base_path + output_report}")
    workbook.save(base_path + output_report)
    print("Excel update completed successfully.")

if __name__ == "__main__":
     
    """
    Entry point for the script. Calls automate_excel_pivoting with the configuration file.
    """
    automate_excel_pivoting('C:\\One Drive\\OneDrive - Tredence\\Desktop\\Agentic_AI\\excel_config.json')
